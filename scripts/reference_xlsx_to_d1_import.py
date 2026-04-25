from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT_XLSX = ROOT / "reference" / "price_minimal_template_数据表_总表.xlsx"
OUTPUT_SQL = ROOT / "tmp" / "reference-xlsx-import.sql"
OUTPUT_SUMMARY = ROOT / "tmp" / "reference-xlsx-import-summary.json"
IMPORT_DATE = date(2026, 4, 25).isoformat()

STORE_MAP = {
    "every 西条御菌店": {
        "id": 7,
        "name": "Gyomu Super Every エブリイ（西条御薗宇店）",
        "chain_brand": "业务超市",
        "location": "西条町御薗宇",
        "note": "大包装和冷冻食品选择多",
    },
    "fresta": {
        "id": 12,
        "name": "FRESTA フレスタ（西条店）",
        "chain_brand": "FRESTA",
        "location": "Saijo 西条",
        "note": "本地常用连锁超市",
    },
    "hallows": {
        "id": 8,
        "name": "HALOWS ハローズ（東広島店）",
        "chain_brand": "HALOWS",
        "location": "Saijo Jike 西条町寺家",
        "note": "营业时间长",
    },
    "lamu": {
        "id": 9,
        "name": "LAMU ラ・ムー（西条店）",
        "chain_brand": "LAMU",
        "location": "Saijo Doyomaru 西条町土与丸",
        "note": "低价导向",
    },
    "wants 西条御菌": {
        "id": 101,
        "name": "Wants ウォンツ（西条御薗宇店）",
        "chain_brand": "Wants",
        "location": "Saijo Misonou 西条町御薗宇",
        "note": "广岛县东广岛市西条町御薗宇2850-1",
    },
    "中央cosmos": {
        "id": 1,
        "name": "Cosmos コスモス（中央店）",
        "chain_brand": "Cosmos",
        "location": "中央",
        "note": "",
    },
    "元气市": {
        "id": 102,
        "name": "とれたて元気市（西条寺家）",
        "chain_brand": "JA",
        "location": "Saijo Jike 西条町寺家",
        "note": "JA交流ひろば「とれたて元気市となりの農家店」 / 广岛县东广岛市西条町寺家7957-1",
    },
    "西条大youme": {
        "id": 4,
        "name": "大youme（東広島）",
        "chain_brand": "youme",
        "location": "Saijo Doyomaru 西条土与丸",
        "note": "中央区域的大youme",
    },
    "骨里香物产店": {
        "id": 103,
        "name": "骨里香熟食物産店（西条中央）",
        "chain_brand": "骨里香",
        "location": "Saijo Chuo 西条中央",
        "note": "广島正宗骨里香熟食店 / 广岛县东广岛市西条中央3-5-20",
    },
}


def main() -> None:
    wb = load_workbook(INPUT_XLSX, data_only=True, read_only=True)
    ws = wb.active
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    idx = {header: i for i, header in enumerate(headers)}

    OUTPUT_SQL.parent.mkdir(parents=True, exist_ok=True)

    product_id_by_key: dict[str, int] = {}
    products: list[dict] = []
    price_records: list[dict] = []
    imported_new_store_keys = {
        key for key, value in STORE_MAP.items() if value["id"] >= 100
    }
    fallback_specs: list[dict] = []
    fallback_dates: list[dict] = []

    next_product_id = 1000
    next_price_record_id = 5000

    for excel_row, row in enumerate(rows[1:], start=2):
        raw_store = clean(row[idx["店铺位置"]])
        if raw_store not in STORE_MAP:
            raise RuntimeError(f"Unmapped store at row {excel_row}: {raw_store}")

        store = STORE_MAP[raw_store]
        name_zh = clean(row[idx["名称(中文)"]])
        name_ja = clean(row[idx["名称(日文)"]])
        barcode = clean_barcode(row[idx["条形码"]])
        product_key = barcode or f"{name_zh.lower()}|{name_ja.lower()}"

        if product_key not in product_id_by_key:
            product_id_by_key[product_key] = next_product_id
            products.append(
                {
                    "id": next_product_id,
                    "name_zh": name_zh,
                    "name_ja": name_ja,
                    "brand": "",
                    "barcode": barcode,
                    "category_id": None,
                    "default_image_url": None,
                    "created_by": "reference-xlsx-import",
                    "created_at": IMPORT_DATE,
                    "updated_at": IMPORT_DATE,
                }
            )
            next_product_id += 1

        normalized_unit = normalize_unit(row[idx["单位"]])
        spec_value, spec_note = normalize_spec(row[idx["规格"]], normalized_unit)
        if spec_note:
            fallback_specs.append(
                {
                    "row": excel_row,
                    "nameZh": name_zh,
                    "store": raw_store,
                    "fallback": spec_note,
                }
            )

        record_date = normalize_date(row[idx["更新时间"]])
        date_note = None
        if record_date == IMPORT_DATE and row[idx["更新时间"]] in (None, ""):
            date_note = f"原更新时间缺失，导入时设为{IMPORT_DATE}"
            fallback_dates.append(
                {"row": excel_row, "nameZh": name_zh, "store": raw_store}
            )

        price_tax_ex = float(row[idx["价格(税前)"]])
        tax_rate = parse_tax_rate(row[idx["税率"]])
        price_tax_in = parse_yen(row[idx["价格(税后)"]])
        unit_price, unit_price_label = calculate_unit_price(
            price_tax_in, spec_value, normalized_unit
        )

        note_parts = []
        for key in ("备注|体验", "总结"):
            value = clean(row[idx[key]])
            if value:
                note_parts.append(value)
        if clean(row[idx["限时特惠？"]]):
            note_parts.append("限时特惠")
        if spec_note:
            note_parts.append(spec_note)
        if date_note:
            note_parts.append(date_note)
        note = " | ".join(note_parts) or None

        price_records.append(
            {
                "id": next_price_record_id,
                "product_id": product_id_by_key[product_key],
                "store_id": store["id"],
                "price_tax_in": price_tax_in,
                "price_tax_ex": price_tax_ex,
                "tax_rate": tax_rate,
                "spec_value": spec_value,
                "unit": normalized_unit,
                "unit_price": unit_price,
                "unit_price_label": unit_price_label,
                "image_url": None,
                "record_date": record_date,
                "note": note,
                "created_by": "reference-xlsx-import",
                "created_at": record_date,
                "updated_at": record_date,
            }
        )
        next_price_record_id += 1

    lines = ["pragma foreign_keys = on;"]

    for raw_key in sorted(imported_new_store_keys, key=lambda x: STORE_MAP[x]["id"]):
        store = STORE_MAP[raw_key]
        lines.append(
            upsert(
                "stores",
                {
                    "id": store["id"],
                    "name": store["name"],
                    "chain_brand": store["chain_brand"],
                    "location": store["location"],
                    "note": store["note"],
                    "created_by": "reference-xlsx-import",
                },
            )
        )

    for product in products:
        lines.append(upsert("products", product))

    for price_record in price_records:
        lines.append(upsert("price_records", price_record))

    lines.append("")
    OUTPUT_SQL.write_text("\n".join(lines), encoding="utf-8")

    summary = {
        "source": str(INPUT_XLSX),
        "rowCount": len(rows) - 1,
        "productCount": len(products),
        "priceRecordCount": len(price_records),
        "newStores": [STORE_MAP[key] for key in sorted(imported_new_store_keys, key=lambda x: STORE_MAP[x]["id"])],
        "fallbackSpecCount": len(fallback_specs),
        "fallbackDateCount": len(fallback_dates),
        "fallbackSpecs": fallback_specs,
        "fallbackDates": fallback_dates,
        "imageStrategy": "ignored",
        "categoryStrategy": "null",
    }
    OUTPUT_SUMMARY.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"Wrote {OUTPUT_SQL}")
    print(f"Wrote {OUTPUT_SUMMARY}")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def normalize_unit(raw_value):
    unit = clean(raw_value)
    if unit == "克":
        return "g"
    if unit == "毫升":
        return "ml"
    if unit == "个":
        return "个"
    raise RuntimeError(f"Unsupported unit: {raw_value}")


def normalize_spec(raw_value, normalized_unit):
    if raw_value not in (None, ""):
        if isinstance(raw_value, (int, float)):
            return float(raw_value), None
        match = re.search(r"\d+(?:\.\d+)?", str(raw_value))
        if match:
            return float(match.group(0)), None
        raise RuntimeError(f"Unsupported spec value: {raw_value}")
    if normalized_unit == "个":
        return 1.0, "规格缺失，按1个兜底"
    if normalized_unit in ("g", "ml"):
        return 100.0, f"规格缺失，按100{normalized_unit}兜底"
    raise RuntimeError(f"Cannot fallback spec for unit: {normalized_unit}")


def normalize_date(raw_value):
    if raw_value in (None, ""):
        return IMPORT_DATE
    if isinstance(raw_value, datetime):
        return raw_value.date().isoformat()
    if isinstance(raw_value, date):
        return raw_value.isoformat()
    text = clean(raw_value)
    return text[:10]


def parse_tax_rate(raw_value):
    return float(clean(raw_value).replace("%", ""))


def parse_yen(raw_value):
    return float(clean(raw_value).replace("¥", "").replace(",", ""))


def calculate_unit_price(price_tax_in, spec_value, normalized_unit):
    if normalized_unit == "g":
        return round(price_tax_in / (spec_value / 100), 4), "/100g"
    if normalized_unit == "ml":
        return round(price_tax_in / (spec_value / 100), 4), "/100ml"
    return round(price_tax_in / spec_value, 4), "/个"


def clean(value):
    return str(value or "").strip()


def clean_barcode(value):
    if value in (None, ""):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def upsert(table, row):
    columns = list(row.keys())
    values = ", ".join(sql_value(row[key]) for key in columns)
    updates = ", ".join(
        f'{quote_ident(key)} = excluded.{quote_ident(key)}'
        for key in columns
        if key != "id"
    )
    return (
        f"insert into {quote_ident(table)} ({', '.join(quote_ident(c) for c in columns)}) "
        f"values ({values}) on conflict(id) do update set {updates};"
    )


def sql_value(value):
    if value is None:
        return "null"
    if isinstance(value, (int, float)):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def quote_ident(name):
    return '"' + str(name).replace('"', '""') + '"'


if __name__ == "__main__":
    main()
